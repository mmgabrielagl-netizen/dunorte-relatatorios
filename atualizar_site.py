#!/usr/bin/env python3
"""
DUNORTE - Script de Atualização Automática do Site
====================================================
Uso: python atualizar_site.py <arquivo>

Aceita:
  - Planilha DRE (.xlsx)     → atualiza o relatório de Resultados no site
  - Relatório de Boletos (.xls / .html) → atualiza o relatório de Boletos no site

Após processar, faz commit e push automático para o GitHub.
"""

import sys, os, base64, re, subprocess, json
from datetime import datetime
from html.parser import HTMLParser

def log(msg): print(f"  ✅ {msg}")
def err(msg): print(f"  ❌ {msg}"); sys.exit(1)

# ── Detecta tipo de arquivo ────────────────────────────────────────────────────
def detect_type(path):
    ext = path.lower().split('.')[-1]
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        preview = f.read(500)
    if ext == 'xlsx': return 'dre'
    if 'BOLETOS' in preview.upper() or 'BOLETO' in preview.upper(): return 'boletos'
    if ext in ('xls','html','htm'): return 'boletos'
    return None

# ── Processa DRE ───────────────────────────────────────────────────────────────
def process_dre(path):
    try:
        import openpyxl
    except ImportError:
        os.system('pip install openpyxl --break-system-packages -q')
        import openpyxl

    log("Lendo planilha DRE...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['DRE']

    # Get month columns
    header = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    month_cols = {}
    for i, v in enumerate(header):
        if v and isinstance(v, str) and v not in ('%', 'Total'):
            month_cols[v] = i

    # Extract all rows
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        label = str(row[1]).strip() if row[1] else None
        if not label: continue
        data[label] = {}
        for month, col in month_cols.items():
            val = row[col] if col < len(row) else None
            data[label][month] = float(val) if val and isinstance(val, (int, float)) else 0.0

    # Key metrics
    meses = list(month_cols.keys())
    receita = data.get('RECEITA', {})
    lucro_liq = data.get('LUCRO LÍQUIDO', data.get('RESULTADO LÍQUIDO', {}))
    custo_direto = data.get('Evento', {})
    comercial = data.get('Comissão Recorrente', {})

    # Build chart data
    chart_receita = [receita.get(m, 0)/1000 for m in meses]
    chart_lucro = [lucro_liq.get(m, 0)/1000 for m in meses]

    # Last month with data
    ultimo_mes = None
    for m in reversed(meses):
        if receita.get(m, 0) != 0:
            ultimo_mes = m
            break

    ul_receita = receita.get(ultimo_mes, 0) if ultimo_mes else 0
    ul_lucro = lucro_liq.get(ultimo_mes, 0) if ultimo_mes else 0
    ul_margem = (ul_lucro / ul_receita * 100) if ul_receita else 0
    ul_custo = custo_direto.get(ultimo_mes, 0) if ultimo_mes else 0

    def fmt_brl(v):
        return f"R$ {v:,.2f}".replace(',','X').replace('.',',').replace('X','.')

    def fmt_k(v):
        if abs(v) >= 1000000: return f"R$ {v/1000000:.1f}M"
        if abs(v) >= 1000: return f"R$ {v/1000:.0f}K"
        return fmt_brl(v)

    log(f"Último mês com dados: {ultimo_mes}")
    log(f"Receita: {fmt_k(ul_receita)}")
    log(f"Lucro Líquido: {fmt_k(ul_lucro)}")
    log(f"Margem: {ul_margem:.1f}%")

    # Build table rows
    grupos = [
        ('RECEITA', ['Contribuição de Associados', 'RECEITA']),
        ('CUSTO DIRETO', ['Indenização Integral', 'Mão De Obra', 'Mão De Obra, Funilaria E Pintura',
                          'Compra De Peças', 'Parabrisas E Vidros Em Geral', 'Carro Reserva',
                          'Guincho', 'Assistência 24Hrs', 'Participações', 'Salvados']),
        ('PESSOAL', ['Salários E Ordenados', 'Vale Transporte', 'Ajuda De Custo', 'Reembolso']),
        ('ADMINISTRATIVO', ['Aluguel', 'Combustível', 'Contabilidade', 'Água']),
        ('COMERCIAL', ['Comissão Recorrente', 'Comissão De Vendas', 'Despesas De Viagem', 'Promoções Comerciais']),
        ('RESULTADO', ['LUCRO LÍQUIDO']),
    ]

    meses_ativos = [m for m in meses if any(data.get(k, {}).get(m, 0) != 0 for k in data)]

    table_rows = ''
    for grupo, itens in grupos:
        is_total = grupo in ('RECEITA', 'RESULTADO')
        bg = '#0B1E35' if grupo == 'RESULTADO' else ('#E8F0FA' if is_total else 'white')
        color = '#F5A800' if grupo == 'RESULTADO' else ('#0B1E35' if is_total else '#333')
        fw = '700' if is_total or grupo == 'RESULTADO' else '400'

        for item in itens:
            row_data = data.get(item, {})
            if not any(row_data.get(m, 0) != 0 for m in meses_ativos): continue
            cells = ''.join(f'<td style="text-align:right;padding:8px 12px;font-size:12px;">{fmt_k(row_data.get(m,0))}</td>' for m in meses_ativos[-6:])
            table_rows += f'''<tr style="background:{bg};color:{color};font-weight:{fw};border-bottom:1px solid #EEF3F9;">
                <td style="padding:8px 12px;font-size:12px;position:sticky;left:0;background:{bg};">{item}</td>{cells}</tr>'''

    month_headers = ''.join(f'<th style="padding:10px 12px;text-align:right;white-space:nowrap;">{m[:3]}</th>' for m in meses_ativos[-6:])

    html = f'''<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>DRE — Dunorte Truck</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet"/>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Inter',sans-serif;background:#F4F6FA;color:#0B1E35;}}
.kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;padding:20px 24px;}}
.kpi{{background:white;border-radius:10px;border:1px solid #D0DCEB;padding:16px 18px;}}
.kpi-label{{font-size:11px;color:#5C7A99;text-transform:uppercase;letter-spacing:.6px;margin-bottom:6px;}}
.kpi-value{{font-size:22px;font-weight:700;color:#0B1E35;}}
.kpi-value.pos{{color:#1A6B3A;}}
.kpi-value.neg{{color:#C0392B;}}
.kpi-sub{{font-size:11px;color:#5C7A99;margin-top:3px;}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:16px;padding:0 24px 20px;}}
.chart-card{{background:white;border-radius:10px;border:1px solid #D0DCEB;padding:16px;}}
.chart-title{{font-size:13px;font-weight:600;margin-bottom:12px;}}
.tbl-wrap{{margin:0 24px 24px;background:white;border-radius:10px;border:1px solid #D0DCEB;overflow:hidden;}}
.tbl-head{{background:#0B1E35;color:white;padding:12px 16px;font-size:13px;font-weight:600;display:flex;align-items:center;justify-content:space-between;}}
.tbl-scroll{{overflow-x:auto;}}
table{{width:100%;border-collapse:collapse;}}
th{{background:#0B1E35;color:#F5A800;padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.6px;font-weight:700;position:sticky;top:0;}}
@media(max-width:768px){{.kpi-grid{{grid-template-columns:1fr 1fr;}}.charts{{grid-template-columns:1fr;}}}}
</style></head><body>

<div class="kpi-grid">
  <div class="kpi">
    <div class="kpi-label">Receita — {ultimo_mes}</div>
    <div class="kpi-value">{fmt_k(ul_receita)}</div>
    <div class="kpi-sub">Contribuição de associados</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Lucro Líquido — {ultimo_mes}</div>
    <div class="kpi-value {'pos' if ul_lucro >= 0 else 'neg'}">{fmt_k(ul_lucro)}</div>
    <div class="kpi-sub">Após todas as despesas</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Margem Líquida</div>
    <div class="kpi-value {'pos' if ul_margem >= 0 else 'neg'}">{ul_margem:.1f}%</div>
    <div class="kpi-sub">Lucro / Receita</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Custo Direto — {ultimo_mes}</div>
    <div class="kpi-value neg">{fmt_k(ul_custo)}</div>
    <div class="kpi-sub">Eventos e sinistros</div>
  </div>
</div>

<div class="charts">
  <div class="chart-card">
    <div class="chart-title">📈 Receita Mensal (R$ mil)</div>
    <canvas id="chartReceita" height="180"></canvas>
  </div>
  <div class="chart-card">
    <div class="chart-title">💰 Lucro Líquido (R$ mil)</div>
    <canvas id="chartLucro" height="180"></canvas>
  </div>
</div>

<div class="tbl-wrap">
  <div class="tbl-head">
    <span>📊 DRE Detalhado</span>
    <span style="font-size:12px;opacity:.7;">Últimos 6 meses</span>
  </div>
  <div class="tbl-scroll">
    <table>
      <thead><tr>
        <th style="min-width:200px;position:sticky;left:0;">Descrição</th>
        {month_headers}
      </tr></thead>
      <tbody>{table_rows}</tbody>
    </table>
  </div>
</div>

<script>
var meses = {json.dumps(meses_ativos[-6:])};
var receita = {json.dumps(chart_receita[-6:])};
var lucro = {json.dumps(chart_lucro[-6:])};

new Chart(document.getElementById('chartReceita'), {{
  type:'bar',
  data:{{labels:meses,datasets:[{{label:'Receita',data:receita,backgroundColor:'rgba(245,168,0,.8)',borderRadius:6}}]}},
  options:{{plugins:{{legend:{{display:false}}}},scales:{{y:{{ticks:{{callback:function(v){{return'R$'+v+'K'}}}}}}}}}}
}});
new Chart(document.getElementById('chartLucro'), {{
  type:'line',
  data:{{labels:meses,datasets:[{{label:'Lucro',data:lucro,borderColor:'#1A6B3A',backgroundColor:'rgba(26,107,58,.1)',fill:true,tension:.4,pointRadius:5}}]}},
  options:{{plugins:{{legend:{{display:false}}}},scales:{{y:{{ticks:{{callback:function(v){{return'R$'+v+'K'}}}}}}}}}}
}});
</script>
</body></html>'''

    return html, 'resultados'

# ── Processa Boletos ───────────────────────────────────────────────────────────
def process_boletos(path):
    log("Lendo relatório de boletos...")

    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()

    class TP(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows=[];self.cr=[];self.cc='';self.inc=False
        def handle_starttag(self,t,a):
            if t=='tr':self.cr=[]
            if t in('td','th'):self.inc=True;self.cc=''
        def handle_endtag(self,t):
            if t in('td','th'):self.cr.append(self.cc.strip());self.inc=False
            if t=='tr' and self.cr:self.rows.append(self.cr[:])
        def handle_data(self,d):
            if self.inc:self.cc+=d

    p=TP()
    p.feed(content)

    data_rows = []
    for r in p.rows:
        if len(r) >= 9 and r[0] and len(r[0]) > 3:
            nome = r[0].strip()
            if nome and not any(x in nome.upper() for x in ['NOME','ASSOCIADO','TOTAL','RELAT','BOLETO']):
                data_rows.append([c.strip() for c in r[:10]])

    log(f"Boletos encontrados: {len(data_rows)}")

    # Stats
    baixados = [r for r in data_rows if 'BAIXADO' in str(r[-1]).upper()]
    abertos = [r for r in data_rows if 'ABERTO' in str(r[-1]).upper() and 'EM ABERTO' not in str(r[-1]).upper()]
    cancelados = [r for r in data_rows if 'CANCEL' in str(r[-1]).upper()]

    rows_json = json.dumps(data_rows[:5000], ensure_ascii=False)
    gerado_em = datetime.now().strftime('%d/%m/%Y às %H:%M')

    html = f'''<!DOCTYPE html><html lang="pt-BR"><head>
<meta charset="UTF-8"/><meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>Relatório de Boletos — Dunorte</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet"/>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Inter',sans-serif;background:#F4F6FA;color:#0B1E35;padding-bottom:40px;}}
.stats{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:16px 20px;}}
.stat{{background:white;border-radius:10px;border:1px solid #D0DCEB;padding:14px 16px;}}
.stat-label{{font-size:11px;color:#5C7A99;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px;}}
.stat-value{{font-size:20px;font-weight:700;}}
.stat-value.total{{color:#0B1E35;}}
.stat-value.baixado{{color:#1A6B3A;}}
.stat-value.aberto{{color:#B85C00;}}
.stat-value.cancelado{{color:#C0392B;}}
.search-bar{{display:flex;gap:10px;align-items:center;margin:0 20px 14px;}}
.search-bar input{{flex:1;padding:9px 14px;border:1.5px solid #D0DCEB;border-radius:8px;font-size:13px;font-family:'Inter',sans-serif;outline:none;background:white;}}
.search-bar input:focus{{border-color:#F5A800;}}
.count{{font-size:12px;color:#5C7A99;white-space:nowrap;font-weight:500;}}
.tbl-wrap{{overflow-x:auto;margin:0 20px;border-radius:10px;box-shadow:0 1px 6px rgba(0,0,0,.07);}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:white;}}
th{{background:#0B1E35;color:#F5A800;padding:10px 12px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.6px;font-weight:700;white-space:nowrap;}}
td{{padding:9px 12px;border-bottom:1px solid #EEF3F9;color:#0B1E35;white-space:nowrap;}}
tr:hover td{{background:#F4F6FA;}}
.sit{{padding:2px 9px;border-radius:12px;font-size:10px;font-weight:700;display:inline-block;}}
.s-b{{background:#E6F5EC;color:#1A6B3A;}}
.s-a{{background:#FEF3E2;color:#B85C00;}}
.s-c{{background:#FDEAEA;color:#C0392B;}}
.gerado{{font-size:11px;color:#5C7A99;padding:8px 20px;}}
@media(max-width:768px){{.stats{{grid-template-columns:1fr 1fr;}}}}
</style></head><body>

<div class="stats">
  <div class="stat"><div class="stat-label">Total de Boletos</div><div class="stat-value total">{len(data_rows)}</div></div>
  <div class="stat"><div class="stat-label">Baixados</div><div class="stat-value baixado">{len(baixados)}</div></div>
  <div class="stat"><div class="stat-label">Em Aberto</div><div class="stat-value aberto">{len(abertos)}</div></div>
  <div class="stat"><div class="stat-label">Cancelados</div><div class="stat-value cancelado">{len(cancelados)}</div></div>
</div>

<div class="search-bar">
  <input type="text" id="srch" placeholder="🔍 Buscar por nome, placa, cooperativa, situação..." oninput="filt()"/>
  <span class="count" id="cnt"></span>
</div>

<div class="tbl-wrap">
  <table>
    <thead><tr>
      <th>#</th><th>Nome</th><th>Placa</th><th>Valor</th><th>Valor Pago</th>
      <th>Emissão</th><th>Pgto</th><th>Vencimento</th><th>Cooperativa</th><th>Voluntário</th><th>Situação</th>
    </tr></thead>
    <tbody id="tbody"></tbody>
  </table>
</div>
<div class="gerado">Gerado em: {gerado_em}</div>

<script>
var all={rows_json};
function sit(s){{
  if(!s)return'';
  var u=s.toUpperCase();
  var cls=u.includes('BAIXADO')?'s-b':u.includes('CANCEL')?'s-c':'s-a';
  return '<span class="sit '+cls+'">'+s+'</span>';
}}
function esc(s){{return s?String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'):''}}
function render(data){{
  document.getElementById('tbody').innerHTML=data.map(function(r,i){{
    return '<tr><td style="color:#5C7A99;">'+(i+1)+'</td><td><strong>'+esc(r[0])+'</strong></td><td style="color:#F5A800;font-weight:600;">'+esc(r[1])+'</td><td>'+esc(r[2])+'</td><td>'+esc(r[3])+'</td><td>'+esc(r[4])+'</td><td>'+esc(r[5])+'</td><td>'+esc(r[6])+'</td><td style="font-size:11px;">'+esc(r[7])+'</td><td style="font-size:11px;">'+esc(r[8])+'</td><td>'+(r[9]?sit(r[9]):'')+'</td></tr>';
  }}).join('');
  document.getElementById('cnt').textContent=data.length+' registros';
}}
function filt(){{
  var q=document.getElementById('srch').value.toLowerCase();
  render(q?all.filter(function(r){{return r.join(' ').toLowerCase().includes(q);}}):all);
}}
render(all);
</script></body></html>'''

    return html, 'boletos'

# ── Atualiza index.html ────────────────────────────────────────────────────────
def update_site(new_html, key, index_path):
    log(f"Lendo index.html atual...")
    with open(index_path, 'r', encoding='utf-8') as f:
        index = f.read()

    # Encode new report as base64
    new_b64 = base64.b64encode(new_html.encode('utf-8')).decode()

    # Find and replace the key in const R = {...}
    pattern = rf"('{key}':')[^']*(')"
    replacement = rf"\g<1>{new_b64}\g<2>"

    new_index = re.sub(pattern, replacement, index)

    if new_index == index:
        err(f"Não encontrei a chave '{key}' no index.html. Verifique se o site está configurado corretamente.")

    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(new_index)

    log(f"index.html atualizado com o relatório '{key}'")
    return True

# ── Git push ───────────────────────────────────────────────────────────────────
def git_push(repo_path, key):
    os.chdir(repo_path)
    now = datetime.now().strftime('%d/%m/%Y %H:%M')
    cmds = [
        ['git', 'add', 'index.html'],
        ['git', 'commit', '-m', f'feat: atualiza relatório {key} — {now}'],
        ['git', 'push']
    ]
    for cmd in cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0:
            err(f"Erro no git: {result.stderr}")
    log("Push para GitHub concluído!")
    log("Site será atualizado na Vercel em ~30 segundos ✅")

# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("\n📋 USO: python atualizar_site.py <caminho_do_arquivo>")
        print("\nExemplos:")
        print("  python atualizar_site.py C:\\Users\\FCJ\\Downloads\\DRE_Junho.xlsx")
        print("  python atualizar_site.py C:\\Users\\FCJ\\Downloads\\boletos.xls")
        sys.exit(0)

    arquivo = sys.argv[1]

    if not os.path.exists(arquivo):
        err(f"Arquivo não encontrado: {arquivo}")

    print(f"\n🚀 Dunorte — Atualização Automática do Site")
    print(f"   Arquivo: {os.path.basename(arquivo)}")
    print()

    tipo = detect_type(arquivo)
    if not tipo:
        err("Tipo de arquivo não reconhecido. Use .xlsx (DRE) ou .xls/.html (Boletos)")

    print(f"  📊 Tipo detectado: {tipo.upper()}")

    if tipo == 'dre':
        html, key = process_dre(arquivo)
    else:
        html, key = process_boletos(arquivo)

    # Find index.html
    script_dir = os.path.dirname(os.path.abspath(__file__))
    index_path = os.path.join(script_dir, 'index.html')

    if not os.path.exists(index_path):
        err(f"index.html não encontrado em {script_dir}")

    update_site(html, key, index_path)
    git_push(script_dir, key)

    print(f"\n🎉 Pronto! O relatório de {tipo.upper()} foi atualizado no site!")
    print(f"   Acesse: https://dunorte-relatatorios.vercel.app")
